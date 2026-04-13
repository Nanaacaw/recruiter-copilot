import os
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from app.core.config import settings


class ExportService:
    def generate_screening_pdf(self, screening: dict, candidate: dict, jd: dict, output_path: str) -> str:
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle("CustomTitle", parent=styles["Title"], fontSize=20, textColor=HexColor("#1E3A5F"))
        heading_style = ParagraphStyle("CustomHeading", parent=styles["Heading2"], fontSize=14, textColor=HexColor("#3B82F6"))
        body_style = ParagraphStyle("CustomBody", parent=styles["Normal"], fontSize=10)

        elements.append(Paragraph("AI Screening Report", title_style))
        elements.append(Spacer(1, 10 * mm))
        elements.append(HRFlowable(width="100%", color=HexColor("#3B82F6")))
        elements.append(Spacer(1, 5 * mm))

        elements.append(Paragraph(f"<b>Candidate:</b> {candidate.get('name', 'N/A')}", body_style))
        elements.append(Paragraph(f"<b>Position:</b> {jd.get('title', 'N/A')}", body_style))
        elements.append(Paragraph(f"<b>Department:</b> {jd.get('department', 'N/A')}", body_style))
        elements.append(Paragraph(f"<b>Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", body_style))
        elements.append(Spacer(1, 10 * mm))

        elements.append(Paragraph("Score Summary", heading_style))
        elements.append(Spacer(1, 3 * mm))

        score_data = [
            ["Criteria", "Score"],
            ["Overall", f"{screening.get('overall_score', 0):.1f}"],
            ["Skills", f"{screening.get('skills_score', 0):.1f}"],
            ["Experience", f"{screening.get('experience_score', 0):.1f}"],
            ["Education", f"{screening.get('education_score', 0):.1f}"],
            ["Certifications", f"{screening.get('certification_score', 0):.1f}"],
        ]
        score_table = Table(score_data, colWidths=[80 * mm, 40 * mm])
        score_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#3B82F6")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(score_table)
        elements.append(Spacer(1, 10 * mm))

        if screening.get("strengths"):
            elements.append(Paragraph("Strengths", heading_style))
            for s in screening["strengths"]:
                elements.append(Paragraph(f"• {s}", body_style))
            elements.append(Spacer(1, 5 * mm))

        if screening.get("weaknesses"):
            elements.append(Paragraph("Weaknesses", heading_style))
            for w in screening["weaknesses"]:
                elements.append(Paragraph(f"• {w}", body_style))
            elements.append(Spacer(1, 5 * mm))

        if screening.get("red_flags"):
            elements.append(Paragraph("Red Flags", heading_style))
            for r in screening["red_flags"]:
                elements.append(Paragraph(f"• {r}", body_style))
            elements.append(Spacer(1, 5 * mm))

        if screening.get("matched_skills"):
            elements.append(Paragraph("Matched Skills", heading_style))
            elements.append(Paragraph(", ".join(screening["matched_skills"]), body_style))
            elements.append(Spacer(1, 5 * mm))

        if screening.get("missing_skills"):
            elements.append(Paragraph("Missing Skills", heading_style))
            elements.append(Paragraph(", ".join(screening["missing_skills"]), body_style))

        doc.build(elements)
        return output_path

    def generate_excel_report(self, screenings: list[dict], output_path: str) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Screening Results"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

        headers = ["Rank", "Candidate", "Email", "Overall Score", "Skills", "Experience", "Education", "Certifications", "Strengths", "Weaknesses"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        sorted_screenings = sorted(screenings, key=lambda x: x.get("overall_score", 0), reverse=True)
        for idx, s in enumerate(sorted_screenings, 1):
            candidate = s.get("candidate", {})
            row = idx + 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=candidate.get("name", ""))
            ws.cell(row=row, column=3, value=candidate.get("email", ""))
            ws.cell(row=row, column=4, value=round(s.get("overall_score", 0), 1))
            ws.cell(row=row, column=5, value=round(s.get("skills_score", 0), 1))
            ws.cell(row=row, column=6, value=round(s.get("experience_score", 0), 1))
            ws.cell(row=row, column=7, value=round(s.get("education_score", 0), 1))
            ws.cell(row=row, column=8, value=round(s.get("certification_score", 0), 1))
            ws.cell(row=row, column=9, value="; ".join(s.get("strengths", [])[:3]))
            ws.cell(row=row, column=10, value="; ".join(s.get("weaknesses", [])[:3]))

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        wb.save(output_path)
        return output_path


export_service = ExportService()
